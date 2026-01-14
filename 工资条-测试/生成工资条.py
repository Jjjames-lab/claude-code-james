#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工资条生成器 - 为每个员工生成独立的Excel工资条文件
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

# 员工工资数据
employees = [
    {"序号": 1, "姓名": "张三", "部门": "管理费", "基本工资": 12000, "绩效工资": 0, "工龄工资": 800, "加班费": 203, "补贴": 10, "应发合计": 1013, "养老保险": 500, "医疗保险": 0, "失业保险": 0, "公积金": 12000, "税前扣除": 1513, "个人所得税": 165.32, "实发工资": 10501.68},
    {"序号": 2, "姓名": "李四", "部门": "管理费", "基本工资": 8500, "绩效工资": 0, "工龄工资": 650, "加班费": 180, "补贴": 8.5, "应发合计": 838.5, "养老保险": 400, "医疗保险": 0, "失业保险": 0, "公积金": 8500, "税前扣除": 1238.5, "个人所得税": 88.45, "实发工资": 7323.05},
    {"序号": 3, "姓名": "王五", "部门": "技术费", "基本工资": 15000, "绩效工资": 0, "工龄工资": 900, "加班费": 250, "补贴": 15, "应发合计": 1165, "养老保险": 600, "医疗保险": 0, "失业保险": 0, "公积金": 15000, "税前扣除": 1765, "个人所得税": 256.78, "实发工资": 12397.22},
    {"序号": 4, "姓名": "赵六", "部门": "管理费", "基本工资": 9800, "绩效工资": 0, "工龄工资": 700, "加班费": 195, "补贴": 9.8, "应发合计": 904.8, "养老保险": 450, "医疗保险": 0, "失业保险": 0, "公积金": 9800, "税前扣除": 1354.8, "个人所得税": 118.92, "实发工资": 8576.28},
    {"序号": 5, "姓名": "陈七", "部门": "技术费", "基本工资": 11500, "绩效工资": 0, "工龄工资": 750, "加班费": 220, "补贴": 11.5, "应发合计": 981.5, "养老保险": 500, "医疗保险": 0, "失业保险": 0, "公积金": 11500, "税前扣除": 1481.5, "个人所得税": 146.88, "实发工资": 9371.62},
    {"序号": 6, "姓名": "刘八", "部门": "管理费", "基本工资": 7200, "绩效工资": 0, "工龄工资": 580, "加班费": 150, "补贴": 7.2, "应发合计": 737.2, "养老保险": 350, "医疗保险": 0, "失业保险": 0, "公积金": 7200, "税前扣除": 1087.2, "个人所得税": 55.68, "实发工资": 6157.12},
    {"序号": 7, "姓名": "杨九", "部门": "技术费", "基本工资": 13200, "绩效工资": 0, "工龄工资": 850, "加班费": 265, "补贴": 13.2, "应发合计": 1128.2, "养老保险": 550, "医疗保险": 0, "失业保险": 0, "公积金": 13200, "税前扣除": 1678.2, "个人所得税": 198.45, "实发工资": 11323.35},
    {"序号": 8, "姓名": "周十", "部门": "管理费", "基本工资": 6800, "绩效工资": 0, "工龄工资": 520, "加班费": 138, "补贴": 6.8, "应发合计": 664.8, "养老保险": 300, "医疗保险": 0, "失业保险": 0, "公积金": 6800, "税前扣除": 964.8, "个人所得税": 42.35, "实发工资": 5842.85},
    {"序号": 9, "姓名": "吴一一", "部门": "技术费", "基本工资": 14500, "绩效工资": 0, "工龄工资": 880, "加班费": 285, "补贴": 14.5, "应发合计": 1179.5, "养老保险": 580, "医疗保险": 0, "失业保险": 0, "公积金": 14500, "税前扣除": 1759.5, "个人所得税": 232.67, "实发工资": 12927.83},
    {"序号": 10, "姓名": "郑一二", "部门": "管理费", "基本工资": 8800, "绩效工资": 0, "工龄工资": 680, "加班费": 175, "补贴": 8.8, "应发合计": 863.8, "养老保险": 420, "医疗保险": 0, "失业保险": 0, "公积金": 8800, "税前扣除": 1283.8, "个人所得税": 95.25, "实发工资": 7620.95},
]

# 表头定义
headers = ["序号", "姓名", "部门", "基本工资", "绩效工资", "工龄工资", "加班费", "补贴",
           "应发合计", "养老保险", "医疗保险", "失业保险", "公积金", "税前扣除", "个人所得税", "实发工资"]

# 边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_salary_slip(employee):
    """为单个员工创建工资条Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "工资条"

    # 设置列宽
    column_widths = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']
    widths = [8, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12]
    for col, width in zip(column_widths, widths):
        ws.column_dimensions[col].width = width

    # 第1行：标题
    ws.merge_cells('A1:Q1')
    title_cell = ws['A1']
    title_cell.value = f"{employee['姓名']}的工资条"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title_cell.font = Font(size=16, bold=True, color='FFFFFF')

    # 第2行：空行
    ws.merge_cells('A2:Q2')

    # 第3行：日期信息
    ws.merge_cells('A3:Q3')
    date_cell = ws['A3']
    date_cell.value = f"发放日期：{datetime.now().strftime('%Y年%m月%d日')}"
    date_cell.font = Font(size=11)
    date_cell.alignment = Alignment(horizontal='right')

    # 第4行：空行
    ws.merge_cells('A4:Q4')

    # 第5行：表头
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = Font(size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border

    # 第6行：员工数据
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = employee[header]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # 高亮实发工资
        if header == "实发工资":
            cell.font = Font(size=11, bold=True, color='008000')
        elif header == "姓名":
            cell.font = Font(size=11, bold=True)

    # 设置行高
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[5].height = 25
    ws.row_dimensions[6].height = 25

    return wb

# 为每个员工生成Excel文件
for emp in employees:
    wb = create_salary_slip(emp)
    filename = f"工资条输出/{emp['姓名']}-工资条.xlsx"
    wb.save(filename)
    print(f"已生成：{filename}")

print("\n✓ 全部完成！共生成 10 个工资条文件")
